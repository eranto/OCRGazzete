from difflib import SequenceMatcher
from functools import reduce
from logging import exception
from numpy import NaN

from rdflib import logging
from Consts import EXCEL_FILE, EXCEL_FILE_RESULTS, PAPERS_FOLDER, SHEET_NAME
import pandas as pd
import openpyxl
import Utils


class ExcelHandler:
    def __init__(self):
        self.excel = pd.read_excel(EXCEL_FILE, SHEET_NAME, engine='openpyxl')
        self.excel_results = openpyxl.load_workbook(EXCEL_FILE_RESULTS)

    def write_to_excel(self, result, paper_date, file_name):
        try:
            worksheet = self.excel_results.worksheets[0]
            num_of_rows = worksheet.max_row+1
            PUBLICATION_DATE_INDEX = 20
            APPLICATION_NUMBER_INDEX = 13
            URL_INDEX = 21
            VERIFIED = 22
            for row_index in range(num_of_rows):  # get publication date
                excel_date = worksheet.cell(
                    row=row_index+1, column=PUBLICATION_DATE_INDEX).value
                if(excel_date is not None and paper_date in excel_date):
                    application_number = str(worksheet.cell(
                        row=row_index+1, column=APPLICATION_NUMBER_INDEX).value)
                    if(application_number in result.keys()):
                        url_to_insert = './'+PAPERS_FOLDER+'/' + \
                            file_name+'/'+application_number+'.png'
                        worksheet.cell(row=row_index+1,
                                       column=URL_INDEX).hyperlink = url_to_insert
                        worksheet.cell(row=row_index+1,
                                       column=URL_INDEX).value = url_to_insert
                        worksheet.cell(row=row_index+1,
                                       column=URL_INDEX).style = "Hyperlink"
                        worksheet.cell(row=row_index+1,
                                       column=VERIFIED).value = 1 if Utils.is_real_file_exist(file_name) else 0
            self.excel_results.save(EXCEL_FILE_RESULTS)
            self.excel_results.close()
        except Exception as e:
            logging.basicConfig(format='%(asctime)s : %(levelname)s : %(message)s',
                                level=logging.INFO, filename=file_name+'.log')
            logging.exception(e)

    # get- date in format d.m.yyyy, returns all rows with this date in Publication dd//mm/yyyy column
    def get_rows_from_date(self, date):
        try:
            lines_from_date = self.excel[self.excel["Publication dd//mm/yyyy"] == 'OG '+date]
            return lines_from_date
        except Exception as e:
            print(
                "Error: cannot extract rows from excel, given date or excel path is is not valid")
            raise e

    def get_row_by_application_number(self, date, num):
        self.rows = self.get_rows_from_date(date)
        return self.rows[self.rows["Application No."] == int(num)]

    def get_class_by_application_number(self, date, num):
        self.rows = self.get_rows_from_date(date)
        return int(self.rows[self.rows["Application No."] == num]["Class No."].values[0])

    def get_trademark_data_by_application_number(self, date, num):
        row = self.get_row_by_application_number(date, num)
        if(row.shape[0] == 1):
            di = {'application_number': str(int(row["Application No."].values[0])),
                  'class_number': row["Class No."].values[0],
                  'type': Utils.parse_trademark_type(row),
                  'initial_no': row["Initial No."].values[0],
                  'date_published': date,
                  'applicant': row["Applicant"].values[0],
                  'local_agent': row["Local Agent"].values[0],
                  'date_applicated': row["Date of Application"].values[0],
                  'sign': row["Sign"].to_list()[0]
                  }
        else:
            raise Exception(f"more than one row were found for app_num {num}",)
        return di

    def is_trademark_type_text(self, app_num):
        row = self.rows[self.rows["Application No."] == app_num]
        return ('Word' in str(row['Symbol Contents']) and str(row['Sign']) != '')

    def get_all_images_app_nums(self, date):
        self.rows = self.get_rows_from_date(date)
        app_nums = []
        for index, row in self.rows.iterrows():
            if('Word' in str(row['Symbol Contents']) and 'Symbol' in str(row['Symbol Contents'])):
                type = 'Image'
            elif(('Word' in str(row['Symbol Contents']) and str(row['Sign']) != '')):
                type = 'Text'
            else:
                type = 'Image'

            if(type == 'Image'):
                app_nums.append(int((row['Application No.'])))
        return app_nums

    def get_all_text_app_nums(self, date):
        self.rows = self.get_rows_from_date(date)
        app_nums = []
        for index, row in self.rows.iterrows():
            if('Word' in str(row['Symbol Contents']) and 'Symbol' in str(row['Symbol Contents'])):
                type = 'Image'
            elif(('Word' in str(row['Symbol Contents']) and str(row['Sign']) != '')):
                type = 'Text'
            else:
                type = 'Image'

            if(type == 'Text'):
                app_nums.append(int((row['Application No.'])))
        return app_nums

    @staticmethod
    def get_rowdata_by_application_number(rows_for_date, num):
        di = None
        for index, row in rows_for_date.iterrows():
            if(str(int(row["Application No."])) == num):
                di = {'application_number': str(int(row["Application No."])),
                      'class_number': str(int(row["Class No."])),
                      'initial_no': row["Initial No."],
                      'applicant': row["Applicant"],
                      'local_agent': row["Local Agent"],
                      'date_applicated': row["Date of Application"],
                      'sign': row["Sign"]
                      }
        return di

    @ staticmethod
    def get_application_numbers_by_class(rows_for_date, class_number):
        li = []
        for index, row in rows_for_date.iterrows():
            if(str(int(row["Class No."])) == class_number):
                li.append(str(int(row["Application No."])))
        return li

    @ staticmethod
    def get_application_numbers_by_class_and_initial(rows_for_date, initial_number, class_number):
        try:
            for index, row in rows_for_date.iterrows():
                if(int(row["Class No."]) == int(class_number) and ((str(row["Initial No."]).isdigit() and int(row["Initial No."]) == int(initial_number)) or (int(row["Application No."]) == int(initial_number)))):
                    return str(int(row["Application No."]))
        except Exception as e:
            logging.exception(e)
            print(e)
            raise e
        return '-1'

    @ staticmethod
    def get_application_numbers_by_application_date(rows_for_date, application_date):
        li = []
        for index, row in rows_for_date.iterrows():
            application_date_from_excel = Utils.get_date_from_text(
                str(row["Date of Application"]))
            if(application_date_from_excel[0] == application_date[0] and application_date_from_excel[1] == application_date[1]):
                li.append(str(int(row["Application No."])))
        return li

    @staticmethod
    def get_countries_from_data_frame(rows_for_date):
        li = []
        for index, row in rows_for_date.iterrows():
            s = str(row["Country of Applicant"])
            # city = str(row["City of Applicant"]).lower().strip() if str(
            #     row["City of Applicant"]) != 'nan' else ''
            s = s.split(',')
            new_s = [el.strip().lower() for el in s]
            new_s.remove('') if '' in new_s else None
            for i, val in enumerate(new_s):
                if(val == 'local'):
                    new_s[i] = 'palestine'
            li.append(new_s)
        return li

    @staticmethod
    def get_cities_from_data_frame(rows_for_date):
        li = []
        for index, row in rows_for_date.iterrows():
            s = str(row["City of Applicant"]) if str(
                row["City of Applicant"]) != 'nan' else None
            if(s == None):
                continue
            # city = str(row["City of Applicant"]).lower().strip() if str(
            #     row["City of Applicant"]) != 'nan' else ''
            s = s.split(',')
            new_s = [el.strip().lower() for el in s]
            new_s.remove('') if '' in new_s else None
            li.append(new_s)
        return li

    @staticmethod
    def get_applicants_from_data_frame(rows_for_date):
        li = []
        for index, row in rows_for_date.iterrows():
            s = str(row["Applicant"]) if str(
                row["Applicant"]) != 'nan' else None
            if(s == None):
                continue
            s = s.split(';')
            new_s = [el.strip().lower() for el in s]
            li.extend(new_s)
        return li

    @staticmethod
    def get_applications_numbers_from_data_frame(rows_for_date):
        li = []
        for index, row in rows_for_date.iterrows():
            s = str(int(row["Application No."])) if str(
                row["Application No."]) != 'nan' else None
            if(s == None):
                continue
            li.append(s)
        return li

    @staticmethod
    def get_application_number_by_country(rows_for_date, countries):
        li = []
        if(len(countries) == 0):
            return []
        for index, row in rows_for_date.iterrows():
            row_countries = str(row["Country of Applicant"]).lower().replace(
                'local', 'palestine')
            country_flag = False
            country_flag = reduce(lambda x, y: x or y, list(
                map(lambda c: c in row_countries, countries)))
            if(country_flag):
                li.append(str(int(row["Application No."])))
        return li

    @staticmethod
    def get_application_number_by_city(rows_for_date, cities):
        li = []
        if(cities is None):
            return []
        for index, row in rows_for_date.iterrows():
            city_flag = reduce(lambda x, y: x or y, list(
                map(lambda c: c in str(row["City of Applicant"]).lower(), cities)))
            if(city_flag):
                li.append(str(int(row["Application No."])))
        return li

    @staticmethod
    def get_application_number_by_applicant(rows_for_date, applicants):
        li = []
        if(applicants is None):
            return []
        for index, row in rows_for_date.iterrows():
            applicant_flag = reduce(lambda x, y: x or y, list(
                map(lambda c: c in str(row["Applicant"]).lower(), applicants)))
            if(applicant_flag):
                li.append(str(int(row["Application No."])))
        return li
